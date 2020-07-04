import tkinter
from tkinter import *
from openpyxl import Workbook
import os
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font
from openpyxl.drawing.image import Image
from tkinter import messagebox
import docx2txt
from PIL import Image as image1, ImageTk
from docx import Document

total = 0

q = [
"1. How many rings on the Olympic flag?",
"2. Who was the first man in space?",
"3. Which company is owned by Bill Gates?",
"4. What kind of animal is a lurcher?",
"5. Who discovered radium?",
"6. What was the capital of Ethiopia?",
"7. What is the largest state in the USA?",
"8. In which city was Alexander Graham Bell born in 1847?",
"9. What digit does not exist in Roman Numerals?",
"10. Which European country is divided into areas called Cantons?",
"11. On which national flag is there an eagle and a snake?",
"12. What is the chemical symbol for tungsten?",
"13. Which bird turns it head upside down to eat?",
"14. In 1969 what category was added to the Nobel prizes?",
"15. Minerva is the Goddess of what?",
"16. Bartommelo Christofori invented what?",
"17. Hibernia was the Roman name for which country?",
"18. What company pioneered floppy discs?",
"19. What animals eye is larger than its brain?",
"20. What is Canada's oldest city founded in 1608?"
]


a0 = ["Five","Six","Four"]
a1 = ["Neil Armstrong","Yuri Gagarin","Buzz Aldrin"]
a2 = ["Apple","Microsoft","Amazon"]
a3 = ["Dog","Cat","Elephant"]
a4 = ["Niels Bohr","Albert Einstein","The Curies"]
a5 = ["Cairo","Dakar‎","Addis Ababa","Rabat‎"]
a6 = ["Arizona","Alaska","California","Hawaii"]
a7 = ["Birmingham","Edinburgh","Cambridge","Liverpool"]
a8 = ["Two","Ten","One","Zero"]
a9 = ["Germany","Sweden","Switzerland", "Italy"]
a10 = ["Germany","USA","Mexico","Russia"]
a11 = ["W","Sb","U","Sn"]
a12 = ["Flamingo","Peacock","Parrot","Stork"]
a13 = ["Chemistry","Economics","Literature","Physics"]
a14 = ["Wisdom","War","Love","Agriculture"]
a15 = ["Guitar","Piano","Violin","Saxophone"]
a16 = ["Norway","Ireland","France","Spain"]
a17 = ["Google","IBM","Lenovo","Intel"]
a18 = ["Lion","Eagle","Ostrich","Whale"]
a19 = ["Toronto","Ottawa","Quebec","Regina"]
a20 = ["Sheep","Cow","Horse","Pig"]




def bnext():

   global windowsQ1
   windowsQ1 = Toplevel(root)
   windowsQ1.title("Question 1")
   windowsQ1.geometry("1920x1080")
   windowsQ1.resizable(0, 0)
   root.withdraw()

   lblSA_space_top0 = Label(windowsQ1, text="", font=("arial", 18, "bold"), width=500, wraplength=400, fg='black').pack(
       side=TOP)
   lblSA = Label(windowsQ1, text="Section A", font=("arial", 20, "bold"), width=500, wraplength=400, fg='black').pack(
       side=TOP)
   lblSA_space_top1 = Label(windowsQ1, text="", font=("arial", 18, "bold"), width=500, wraplength=400, fg='black').pack(
       side=TOP)
   lblSA1 = Label(windowsQ1, text="Please Answer all 25 questions; ", font=("arial", 18, "bold"), width=700, wraplength=1200,
                 fg='black').pack(side=TOP)
   lblSA2 = Label(windowsQ1, text="for each question, choose the correct answer.", font=("arial", 18, "bold"), width=700,
                  wraplength=1200, fg='black').pack(side=TOP)

   lblSA_space_top2 = Label(windowsQ1, text="", font=("arial", 18, "bold"), width=500, wraplength=400, fg='black').pack(
       side=TOP)

   lblQ1_space1 = Label(windowsQ1, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)
   lblQ1 = Label(windowsQ1,text = q[0],justify="center", font = ('arial',18,'bold')).pack(side = TOP)
   lblQ1_space2 = Label(windowsQ1, text=" ", justify="center", font=('arial', 15, 'bold')).pack(side=TOP)

   cbQ11 = Radiobutton(windowsQ1, text=a0[0], font=("Times", 16),value=0,variable = v0,command = checked).pack(pady=5)
   cbQ12 = Radiobutton(windowsQ1, text=a0[1], font=("Times", 16),value=1,variable = v0,command = checked).pack(pady=5)
   cbQ13 = Radiobutton(windowsQ1, text=a0[2], font=("Times", 16),value=2,variable = v0,command = checked).pack(pady=5)

   btnQ11 = Button(windowsQ1,text = "Next",font = ('arial',18,'bold'),fg = 'blue',  command = bnext2).pack(side=RIGHT, padx=205)
   btnQ12 = Button(windowsQ1,text = "Back",font = ('arial',18,'bold'),fg = 'blue', command = bback).pack(side=LEFT, padx=205)

   windowsQ1.mainloop()



def bnext2():

   global windowsQ2
   windowsQ2 = Toplevel(windowsQ1)
   windowsQ2.title("Question 2")
   windowsQ2.geometry("1920x1080")
   windowsQ2.resizable(0, 0)
   windowsQ1.withdraw()

   lblQ2_space1 = Label(windowsQ2, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)
   lblQ2_space11 = Label(windowsQ2, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)
   lblQ2 = Label(windowsQ2, text=q[1], font=('arial', 18, 'bold')).pack(side=TOP)
   lblQ2_space2 = Label(windowsQ2, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)

   cbQ21 = Radiobutton(windowsQ2, text=a1[0], font=("Times", 16),value=0,variable = v1,command = checked).pack(pady=5)
   cbQ22 = Radiobutton(windowsQ2, text=a1[1], font=("Times", 16),value=1,variable = v1,command = checked).pack(pady=5)
   cbQ23 = Radiobutton(windowsQ2, text=a1[2], font=("Times", 16),value=2,variable = v1,command = checked).pack(pady=5)

   btnQ21 = Button(windowsQ2,text = "Next",font = ('arial',18,'bold'),fg = 'blue',command = bnext3).pack(side=RIGHT, padx=205)
   btnQ22 = Button(windowsQ2,text = "Back",font = ('arial',18,'bold'),fg = 'blue',command = bback2).pack(side=LEFT, padx=205)
   windowsQ2.mainloop()


def bnext3():

   global windowsQ3
   windowsQ3 = Toplevel(windowsQ2)
   windowsQ3.title("Question 3")
   windowsQ3.geometry("1920x1080")
   windowsQ3.resizable(0, 0)
   windowsQ2.withdraw()
   lblQ3_space1 = Label(windowsQ3, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)
   lblQ3_space11 = Label(windowsQ3, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)
   lblQ3 = Label(windowsQ3, text=q[2], font=('arial', 18, 'bold')).pack(side=TOP)
   lblQ3_space2 = Label(windowsQ3, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)

   cbQ31 = Radiobutton(windowsQ3, text=a2[0], font=("Times", 16), value=0,variable = v2,command = checked).pack(pady=5)
   cbQ32 = Radiobutton(windowsQ3, text=a2[1], font=("Times", 16),value=1,variable = v2,command = checked).pack(pady=5)
   cbQ33 = Radiobutton(windowsQ3, text=a2[2], font=("Times", 16),value=2,variable = v2,command = checked).pack(pady=5)

   btnQ31 = Button(windowsQ3,text = "Next",font = ('arial',18,'bold'),fg = 'blue',command = bnext4).pack(side=RIGHT, padx=205)
   btnQ32 = Button(windowsQ3,text = "Back",font = ('arial',18,'bold'),fg = 'blue',command = bback3).pack(side=LEFT, padx=205)
   windowsQ3.mainloop()




def bnext4():

   global windowsQ4
   windowsQ4 = Toplevel(windowsQ3)
   windowsQ4.title("Question 4")
   windowsQ4.geometry("1920x1080")
   windowsQ4.resizable(0, 0)
   windowsQ3.withdraw()
   lblQ4_space1 = Label(windowsQ4, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)
   lblQ4_space11 = Label(windowsQ4, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)
   lblQ4 = Label(windowsQ4, text=q[3], font=('arial', 18, 'bold')).pack(side=TOP)
   lblQ4_space2 = Label(windowsQ4, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)

   cbQ41 = Radiobutton(windowsQ4, text=a3[0], font=("Times", 16), value=0,variable = v3,command = checked).pack(pady=5)
   cbQ42 = Radiobutton(windowsQ4, text=a3[1], font=("Times", 16),value=1,variable = v3,command = checked).pack(pady=5)
   cbQ43 = Radiobutton(windowsQ4, text=a3[2], font=("Times", 16),value=2,variable = v3,command = checked).pack(pady=5)

   btnQ41 = Button(windowsQ4,text = "Next",font = ('arial',18,'bold'),fg = 'blue',command = bnext5).pack(side=RIGHT, padx=205)
   btnQ42 = Button(windowsQ4,text = "Back",font = ('arial',18,'bold'),fg = 'blue',command = bback4).pack(side=LEFT, padx=205)
   windowsQ4.mainloop()




def bnext5():

   global windowsQ5
   windowsQ5 = Toplevel(windowsQ4)
   windowsQ5.title("Question 5")
   windowsQ5.geometry("1920x1080")
   windowsQ5.resizable(0, 0)
   windowsQ4.withdraw()
   lblQ5_space1 = Label(windowsQ5, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)
   lblQ5_space11 = Label(windowsQ5, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)
   lblQ5 = Label(windowsQ5, text=q[4], font=('arial', 18, 'bold')).pack(side=TOP)
   lblQ5_space2 = Label(windowsQ5, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)

   cbQ51 = Radiobutton(windowsQ5, text=a4[0], font=("Times", 16), value=0,variable = v4,command = checked).pack(pady=5)
   cbQ52 = Radiobutton(windowsQ5, text=a4[1], font=("Times", 16),value=1,variable = v4,command = checked).pack(pady=5)
   cbQ53 = Radiobutton(windowsQ5, text=a4[2], font=("Times", 16),value=2,variable = v4,command = checked).pack(pady=5)

   btnQ51 = Button(windowsQ5,text = "Next",font = ('arial',18,'bold'),fg = 'blue',command = bnext6).pack(side=RIGHT, padx=205)
   btnQ52 = Button(windowsQ5,text = "Back",font = ('arial',18,'bold'),fg = 'blue',command = bback5).pack(side=LEFT, padx=205)
   windowsQ5.mainloop()



def bnext6():

   global windowsQ6
   windowsQ6 = Toplevel(windowsQ5)
   windowsQ6.title("Question 6")
   windowsQ6.geometry("1920x1080")
   windowsQ6.resizable(0, 0)
   windowsQ5.withdraw()
   lblQ6_space1 = Label(windowsQ6, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)
   lblQ6_space11 = Label(windowsQ6, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)
   lblQ6 = Label(windowsQ6, text=q[5], font=('arial', 18, 'bold')).pack(side=TOP)
   lblQ6_space2 = Label(windowsQ6, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)

   cbQ61 = Radiobutton(windowsQ6, text=a5[0], font=("Times", 16), value=0,variable = v5,command = checked).pack(pady=5)
   cbQ62 = Radiobutton(windowsQ6, text=a5[1], font=("Times", 16),value=1,variable = v5,command = checked).pack(pady=5)
   cbQ63 = Radiobutton(windowsQ6, text=a5[2], font=("Times", 16),value=2,variable = v5,command = checked).pack(pady=5)
   cbQ64 = Radiobutton(windowsQ6, text=a5[3], font=("Times", 16), value=3, variable=v5, command=checked).pack(pady=5)

   btnQ61 = Button(windowsQ6,text = "Next",font = ('arial',18,'bold'),fg = 'blue',command = bnext7).pack(side=RIGHT, padx=205)
   btnQ62 = Button(windowsQ6,text = "Back",font = ('arial',18,'bold'),fg = 'blue',command = bback6).pack(side=LEFT, padx=205)
   windowsQ6.mainloop()



def bnext7():

   global windowsQ7
   windowsQ7 = Toplevel(windowsQ6)
   windowsQ7.title("Question 7")
   windowsQ7.geometry("1920x1080")
   windowsQ7.resizable(0, 0)
   windowsQ6.withdraw()
   lblQ7_space1 = Label(windowsQ7, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)
   lblQ7_space11 = Label(windowsQ7, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)
   lblQ7 = Label(windowsQ7, text=q[6], font=('arial', 18, 'bold')).pack(side=TOP)
   lblQ7_space2 = Label(windowsQ7, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)

   cbQ71 = Radiobutton(windowsQ7, text=a6[0], font=("Times", 16), value=0,variable = v6,command = checked).pack(pady=5)
   cbQ72 = Radiobutton(windowsQ7, text=a6[1], font=("Times", 16),value=1,variable = v6,command = checked).pack(pady=5)
   cbQ73 = Radiobutton(windowsQ7, text=a6[2], font=("Times", 16),value=2,variable = v6,command = checked).pack(pady=5)
   cbQ74 = Radiobutton(windowsQ7, text=a6[3], font=("Times", 16),value=3,variable = v6,command = checked).pack(pady=5)

   btnQ71 = Button(windowsQ7,text = "Next",font = ('arial',18,'bold'),fg = 'blue',command = bnext8).pack(side=RIGHT, padx=205)
   btnQ72 = Button(windowsQ7,text = "Back",font = ('arial',18,'bold'),fg = 'blue',command = bback7).pack(side=LEFT, padx=205)
   windowsQ7.mainloop()



def bnext8():

   global windowsQ8
   windowsQ8 = Toplevel(windowsQ7)
   windowsQ8.title("Question 8")
   windowsQ8.geometry("1920x1080")
   windowsQ8.resizable(0, 0)
   windowsQ7.withdraw()
   lblQ8_space1 = Label(windowsQ8, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)
   lblQ8_space11 = Label(windowsQ8, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)
   lblQ8 = Label(windowsQ8, text=q[7], font=('arial', 18, 'bold')).pack(side=TOP)
   lblQ8_space2 = Label(windowsQ8, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)

   cbQ81 = Radiobutton(windowsQ8, text=a7[0], font=("Times", 16), value=0,variable = v7,command = checked).pack(pady=5)
   cbQ82 = Radiobutton(windowsQ8, text=a7[1], font=("Times", 16),value=1,variable = v7,command = checked).pack(pady=5)
   cbQ83 = Radiobutton(windowsQ8, text=a7[2], font=("Times", 16),value=2,variable = v7,command = checked).pack(pady=5)
   cbQ84 = Radiobutton(windowsQ8, text=a7[3], font=("Times", 16),value=3,variable = v7,command = checked).pack(pady=5)

   btnQ81 = Button(windowsQ8,text = "Next",font = ('arial',18,'bold'),fg = 'blue',command = bnext9).pack(side=RIGHT, padx=205)
   btnQ82 = Button(windowsQ8,text = "Back",font = ('arial',18,'bold'),fg = 'blue',command = bback8).pack(side=LEFT, padx=205)
   windowsQ8.mainloop()



def bnext9():

   global windowsQ9
   windowsQ9 = Toplevel(windowsQ8)
   windowsQ9.title("Question 9")
   windowsQ9.geometry("1920x1080")
   windowsQ9.resizable(0, 0)
   windowsQ8.withdraw()
   lblQ9_space1 = Label(windowsQ9, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)
   lblQ9_space11 = Label(windowsQ9, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)
   lblQ9 = Label(windowsQ9, text=q[8], font=('arial', 18, 'bold')).pack(side=TOP)
   lblQ9_space2 = Label(windowsQ9, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)

   cbQ91 = Radiobutton(windowsQ9, text=a8[0], font=("Times", 16), value=0,variable = v8,command = checked).pack(pady=5)
   cbQ92 = Radiobutton(windowsQ9, text=a8[1], font=("Times", 16),value=1,variable = v8,command = checked).pack(pady=5)
   cbQ93 = Radiobutton(windowsQ9, text=a8[2], font=("Times", 16),value=2,variable = v8,command = checked).pack(pady=5)
   cbQ94 = Radiobutton(windowsQ9, text=a8[3], font=("Times", 16),value=3,variable = v8,command = checked).pack(pady=5)

   btnQ91 = Button(windowsQ9,text = "Next",font = ('arial',18,'bold'),fg = 'blue',command = bnext10).pack(side=RIGHT, padx=205)
   btnQ92 = Button(windowsQ9,text = "Back",font = ('arial',18,'bold'),fg = 'blue',command = bback9).pack(side=LEFT, padx=205)
   windowsQ9.mainloop()


def bnext10():

   global windowsQ10
   windowsQ10 = Toplevel(windowsQ9)
   windowsQ10.title("Question 10")
   windowsQ10.geometry("1920x1080")
   windowsQ10.resizable(0, 0)
   windowsQ9.withdraw()
   lblQ9_space1 = Label(windowsQ10, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)
   lblQ9_space11 = Label(windowsQ10, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)
   lblQ9 = Label(windowsQ10, text=q[9], font=('arial', 18, 'bold')).pack(side=TOP)
   lblQ9_space2 = Label(windowsQ10, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)

   cbQ91 = Radiobutton(windowsQ10, text=a9[0], font=("Times", 16), value=0,variable = v9,command = checked).pack(pady=5)
   cbQ92 = Radiobutton(windowsQ10, text=a9[1], font=("Times", 16),value=1,variable = v9,command = checked).pack(pady=5)
   cbQ93 = Radiobutton(windowsQ10, text=a9[2], font=("Times", 16),value=2,variable = v9,command = checked).pack(pady=5)
   cbQ94 = Radiobutton(windowsQ10, text=a9[3], font=("Times", 16),value=3,variable = v9,command = checked).pack(pady=5)

   btnQ91 = Button(windowsQ10,text = "Next",font = ('arial',18,'bold'),fg = 'blue',command = bnext11).pack(side=RIGHT, padx=205)
   btnQ92 = Button(windowsQ10,text = "Back",font = ('arial',18,'bold'),fg = 'blue',command = bback10).pack(side=LEFT, padx=205)
   windowsQ10.mainloop()


def bnext11():

   global windowsQ11
   windowsQ11 = Toplevel(windowsQ10)
   windowsQ11.title("Question 11")
   windowsQ11.geometry("1920x1080")
   windowsQ11.resizable(0, 0)
   windowsQ10.withdraw()
   lblQ11_space1 = Label(windowsQ11, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)
   lblQ11_space11 = Label(windowsQ11, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)
   lblQ11 = Label(windowsQ11, text=q[10], font=('arial', 18, 'bold')).pack(side=TOP)
   lblQ11_space2 = Label(windowsQ11, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)

   cbQ111 = Radiobutton(windowsQ11, text=a10[0], font=("Times", 16), value=0,variable = v10,command = checked).pack(pady=5)
   cbQ112 = Radiobutton(windowsQ11, text=a10[1], font=("Times", 16),value=1,variable = v10,command = checked).pack(pady=5)
   cbQ113 = Radiobutton(windowsQ11, text=a10[2], font=("Times", 16),value=2,variable = v10,command = checked).pack(pady=5)
   cbQ114 = Radiobutton(windowsQ11, text=a10[3], font=("Times", 16),value=3,variable = v10,command = checked).pack(pady=5)

   btnQ111 = Button(windowsQ11,text = "Next",font = ('arial',18,'bold'),fg = 'blue',command = bnext12).pack(side=RIGHT, padx=205)
   btnQ112 = Button(windowsQ11,text = "Back",font = ('arial',18,'bold'),fg = 'blue',command = bback11).pack(side=LEFT, padx=205)
   windowsQ11.mainloop()


def bnext12():

   global windowsQ12
   windowsQ12 = Toplevel(windowsQ11)
   windowsQ12.title("Question 12")
   windowsQ12.geometry("1920x1080")
   windowsQ12.resizable(0, 0)
   windowsQ11.withdraw()
   lblQ12_space1 = Label(windowsQ12, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)
   lblQ12_space11 = Label(windowsQ12, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)
   lblQ12 = Label(windowsQ12, text=q[11], font=('arial', 18, 'bold')).pack(side=TOP)
   lblQ12_space2 = Label(windowsQ12, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)

   cbQ121 = Radiobutton(windowsQ12, text=a11[0], font=("Times", 16), value=0,variable = v11,command = checked).pack(pady=5)
   cbQ122 = Radiobutton(windowsQ12, text=a11[1], font=("Times", 16),value=1,variable = v11,command = checked).pack(pady=5)
   cbQ123 = Radiobutton(windowsQ12, text=a11[2], font=("Times", 16),value=2,variable = v11,command = checked).pack(pady=5)
   cbQ124 = Radiobutton(windowsQ12, text=a11[3], font=("Times", 16),value=3,variable = v11,command = checked).pack(pady=5)

   btnQ121 = Button(windowsQ12,text = "Next",font = ('arial',18,'bold'),fg = 'blue',command = bnext13).pack(side=RIGHT, padx=205)
   btnQ122 = Button(windowsQ12,text = "Back",font = ('arial',18,'bold'),fg = 'blue',command = bback12).pack(side=LEFT, padx=205)
   windowsQ12.mainloop()


def bnext13():

   global windowsQ13
   windowsQ13 = Toplevel(windowsQ12)
   windowsQ13.title("Question 13")
   windowsQ13.geometry("1920x1080")
   windowsQ13.resizable(0, 0)
   windowsQ12.withdraw()
   lblQ13_space1 = Label(windowsQ13, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)
   lblQ13_space11 = Label(windowsQ13, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)
   lblQ13 = Label(windowsQ13, text=q[12], font=('arial', 18, 'bold')).pack(side=TOP)
   lblQ13_space2 = Label(windowsQ13, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)

   cbQ131 = Radiobutton(windowsQ13, text=a12[0], font=("Times", 16), value=0,variable = v12,command = checked).pack(pady=5)
   cbQ132 = Radiobutton(windowsQ13, text=a12[1], font=("Times", 16),value=1,variable = v12,command = checked).pack(pady=5)
   cbQ133 = Radiobutton(windowsQ13, text=a12[2], font=("Times", 16),value=2,variable = v12,command = checked).pack(pady=5)
   cbQ134 = Radiobutton(windowsQ13, text=a12[3], font=("Times", 16),value=3,variable = v12,command = checked).pack(pady=5)

   btnQ131 = Button(windowsQ13,text = "Next",font = ('arial',18,'bold'),fg = 'blue',command = bnext14).pack(side=RIGHT, padx=205)
   btnQ132 = Button(windowsQ13,text = "Back",font = ('arial',18,'bold'),fg = 'blue',command = bback13).pack(side=LEFT, padx=205)
   windowsQ13.mainloop()


def bnext14():

   global windowsQ14
   windowsQ14 = Toplevel(windowsQ13)
   windowsQ14.title("Question 14")
   windowsQ14.geometry("1920x1080")
   windowsQ14.resizable(0, 0)
   windowsQ13.withdraw()
   lblQ14_space1 = Label(windowsQ14, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)
   lblQ14_space11 = Label(windowsQ14, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)
   lblQ14 = Label(windowsQ14, text=q[13], font=('arial', 18, 'bold')).pack(side=TOP)
   lblQ14_space2 = Label(windowsQ14, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)

   cbQ141 = Radiobutton(windowsQ14, text=a13[0], font=("Times", 16), value=0,variable = v13,command = checked).pack(pady=5)
   cbQ142 = Radiobutton(windowsQ14, text=a13[1], font=("Times", 16),value=1,variable = v13,command = checked).pack(pady=5)
   cbQ143 = Radiobutton(windowsQ14, text=a13[2], font=("Times", 16),value=2,variable = v13,command = checked).pack(pady=5)
   cbQ144 = Radiobutton(windowsQ14, text=a13[3], font=("Times", 16),value=3,variable = v13,command = checked).pack(pady=5)

   btnQ141 = Button(windowsQ14,text = "Next",font = ('arial',18,'bold'),fg = 'blue',command = bnext15).pack(side=RIGHT, padx=205)
   btnQ142 = Button(windowsQ14,text = "Back",font = ('arial',18,'bold'),fg = 'blue',command = bback14).pack(side=LEFT, padx=205)
   windowsQ14.mainloop()


def bnext15():

   global windowsQ15
   windowsQ15 = Toplevel(windowsQ14)
   windowsQ15.title("Question 15")
   windowsQ15.geometry("1920x1080")
   windowsQ15.resizable(0, 0)
   windowsQ14.withdraw()
   lblQ15_space1 = Label(windowsQ15, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)
   lblQ15_space11 = Label(windowsQ15, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)
   lblQ15 = Label(windowsQ15, text=q[14], font=('arial', 18, 'bold')).pack(side=TOP)
   lblQ15_space2 = Label(windowsQ15, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)

   cbQ151 = Radiobutton(windowsQ15, text=a14[0], font=("Times", 16), value=0,variable = v14,command = checked).pack(pady=5)
   cbQ152 = Radiobutton(windowsQ15, text=a14[1], font=("Times", 16),value=1,variable = v14,command = checked).pack(pady=5)
   cbQ153 = Radiobutton(windowsQ15, text=a14[2], font=("Times", 16),value=2,variable = v14,command = checked).pack(pady=5)
   cbQ154 = Radiobutton(windowsQ15, text=a14[3], font=("Times", 16),value=3,variable = v14,command = checked).pack(pady=5)

   btnQ151 = Button(windowsQ15,text = "Next",font = ('arial',18,'bold'),fg = 'blue',command = bnext16).pack(side=RIGHT, padx=205)
   btnQ152 = Button(windowsQ15,text = "Back",font = ('arial',18,'bold'),fg = 'blue',command = bback15).pack(side=LEFT, padx=205)
   windowsQ15.mainloop()



def bnext16():

   global windowsQ16
   windowsQ16 = Toplevel(windowsQ15)
   windowsQ16.title("Question 16")
   windowsQ16.geometry("1920x1080")
   windowsQ16.resizable(0, 0)
   windowsQ15.withdraw()
   lblQ16_space1 = Label(windowsQ16, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)
   lblQ16_space11 = Label(windowsQ16, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)
   lblQ16 = Label(windowsQ16, text=q[15], font=('arial', 18, 'bold')).pack(side=TOP)
   lblQ16_space2 = Label(windowsQ16, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)

   cbQ161 = Radiobutton(windowsQ16, text=a15[0], font=("Times", 16), value=0,variable = v15,command = checked).pack(pady=5)
   cbQ162 = Radiobutton(windowsQ16, text=a15[1], font=("Times", 16),value=1,variable = v15,command = checked).pack(pady=5)
   cbQ163 = Radiobutton(windowsQ16, text=a15[2], font=("Times", 16),value=2,variable = v15,command = checked).pack(pady=5)
   cbQ164 = Radiobutton(windowsQ16, text=a15[3], font=("Times", 16),value=3,variable = v15,command = checked).pack(pady=5)

   btnQ161 = Button(windowsQ16,text = "Next",font = ('arial',18,'bold'),fg = 'blue',command = bnext17).pack(side=RIGHT, padx=205)
   btnQ162 = Button(windowsQ16,text = "Back",font = ('arial',18,'bold'),fg = 'blue',command = bback16).pack(side=LEFT, padx=205)
   windowsQ16.mainloop()



def bnext17():

   global windowsQ17
   windowsQ17 = Toplevel(windowsQ16)
   windowsQ17.title("Question 17")
   windowsQ17.geometry("1920x1080")
   windowsQ17.resizable(0, 0)
   windowsQ16.withdraw()
   lblQ17_space1 = Label(windowsQ17, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)
   lblQ17_space11 = Label(windowsQ17, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)
   lblQ17 = Label(windowsQ17, text=q[16], font=('arial', 18, 'bold')).pack(side=TOP)
   lblQ17_space2 = Label(windowsQ17, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)

   cbQ171 = Radiobutton(windowsQ17, text=a16[0], font=("Times", 16), value=0,variable = v16,command = checked).pack(pady=5)
   cbQ172 = Radiobutton(windowsQ17, text=a16[1], font=("Times", 16),value=1,variable = v16,command = checked).pack(pady=5)
   cbQ173 = Radiobutton(windowsQ17, text=a16[2], font=("Times", 16),value=2,variable = v16,command = checked).pack(pady=5)
   cbQ174 = Radiobutton(windowsQ17, text=a16[3], font=("Times", 16),value=3,variable = v16,command = checked).pack(pady=5)

   btnQ171 = Button(windowsQ17,text = "Next",font = ('arial',18,'bold'),fg = 'blue',command = bnext18).pack(side=RIGHT, padx=205)
   btnQ172 = Button(windowsQ17,text = "Back",font = ('arial',18,'bold'),fg = 'blue',command = bback17).pack(side=LEFT, padx=205)
   windowsQ17.mainloop()



def bnext18():

   global windowsQ18
   windowsQ18 = Toplevel(windowsQ17)
   windowsQ18.title("Question 18")
   windowsQ18.geometry("1920x1080")
   windowsQ18.resizable(0, 0)
   windowsQ17.withdraw()
   lblQ18_space1 = Label(windowsQ18, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)
   lblQ18_space11 = Label(windowsQ18, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)
   lblQ18 = Label(windowsQ18, text=q[17], font=('arial', 18, 'bold')).pack(side=TOP)
   lblQ18_space2 = Label(windowsQ18, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)

   cbQ181 = Radiobutton(windowsQ18, text=a17[0], font=("Times", 16), value=0,variable = v17,command = checked).pack(pady=5)
   cbQ182 = Radiobutton(windowsQ18, text=a17[1], font=("Times", 16),value=1,variable = v17,command = checked).pack(pady=5)
   cbQ183 = Radiobutton(windowsQ18, text=a17[2], font=("Times", 16),value=2,variable = v17,command = checked).pack(pady=5)
   cbQ184 = Radiobutton(windowsQ18, text=a17[3], font=("Times", 16),value=3,variable = v17,command = checked).pack(pady=5)

   btnQ181 = Button(windowsQ18,text = "Next",font = ('arial',18,'bold'),fg = 'blue',command = bnext19).pack(side=RIGHT, padx=205)
   btnQ182 = Button(windowsQ18,text = "Back",font = ('arial',18,'bold'),fg = 'blue',command = bback18).pack(side=LEFT, padx=205)
   windowsQ18.mainloop()



def bnext19():

   global windowsQ19
   windowsQ19 = Toplevel(windowsQ18)
   windowsQ19.title("Question 19")
   windowsQ19.geometry("1920x1080")
   windowsQ19.resizable(0, 0)
   windowsQ18.withdraw()
   lblQ19_space1 = Label(windowsQ19, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)
   lblQ19_space11 = Label(windowsQ19, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)
   lblQ19 = Label(windowsQ19, text=q[18], font=('arial', 18, 'bold')).pack(side=TOP)
   lblQ19_space2 = Label(windowsQ19, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)

   cbQ191 = Radiobutton(windowsQ19, text=a18[0], font=("Times", 16), value=0,variable = v18,command = checked).pack(pady=5)
   cbQ192 = Radiobutton(windowsQ19, text=a18[1], font=("Times", 16),value=1,variable = v18,command = checked).pack(pady=5)
   cbQ193 = Radiobutton(windowsQ19, text=a18[2], font=("Times", 16),value=2,variable = v18,command = checked).pack(pady=5)
   cbQ194 = Radiobutton(windowsQ19, text=a18[3], font=("Times", 16),value=3,variable = v18,command = checked).pack(pady=5)

   btnQ191 = Button(windowsQ19,text = "Next",font = ('arial',18,'bold'),fg = 'blue',command = bnext20).pack(side=RIGHT, padx=205)
   btnQ192 = Button(windowsQ19,text = "Back",font = ('arial',18,'bold'),fg = 'blue',command = bback19).pack(side=LEFT, padx=205)
   windowsQ19.mainloop()



def bnext20():

   global windowsQ20
   windowsQ20 = Toplevel(windowsQ19)
   windowsQ20.title("Question 20")
   windowsQ20.geometry("1920x1080")
   windowsQ20.resizable(0, 0)
   windowsQ19.withdraw()
   lblQ20_space1 = Label(windowsQ20, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)
   lblQ20_space11 = Label(windowsQ20, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)
   lblQ20 = Label(windowsQ20, text=q[19], font=('arial', 18, 'bold')).pack(side=TOP)
   lblQ20_space2 = Label(windowsQ20, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)

   cbQ201 = Radiobutton(windowsQ20, text=a19[0], font=("Times", 16), value=0,variable = v19,command = checked).pack(pady=5)
   cbQ202 = Radiobutton(windowsQ20, text=a19[1], font=("Times", 16),value=1,variable = v19,command = checked).pack(pady=5)
   cbQ203 = Radiobutton(windowsQ20, text=a19[2], font=("Times", 16),value=2,variable = v19,command = checked).pack(pady=5)
   cbQ204 = Radiobutton(windowsQ20, text=a19[3], font=("Times", 16),value=3,variable = v19,command = checked).pack(pady=5)

   btnQ201 = Button(windowsQ20,text = "Next",font = ('arial',18,'bold'),fg = 'blue',command = confirmFinish).pack(side=RIGHT, padx=205)
   btnQ202 = Button(windowsQ20,text = "Back",font = ('arial',18,'bold'),fg = 'blue',command = bback20).pack(side=LEFT, padx=205)
   windowsQ20.mainloop()



def bnext21():

   global windows21
   windows21 = Toplevel(windowsQ20)
   windows21.title("Final Result")
   windows21.geometry("1920x1080")
   windowsQ20.withdraw()
   bquit = Button(windows21, text="Close", font=('arial', 18, 'bold'),
                  fg='blue', command=windows21.destroy).pack(side=BOTTOM, pady=250)
   lbl30_space1 = Label(windows21, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)
   lbl30_space2 = Label(windows21, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)
   lbl30_space3 = Label(windows21, text=" ", justify="center", font = ('arial',22,'bold')).pack(side=TOP)
   if c >= 14:
       lblPass = Label(windows21, text="Congratulations!, You have passed the quiz!", font=("arial", 20, "bold"), fg="green")
       lblPass.pack(side=TOP)
   else:
       lblFail = Label(windows21, text="You have failed!", font=("arial", 20, "bold"), fg="red")
       lblFail.pack(side=TOP)

   windows21.mainloop()



def bback():
    windowsQ1.withdraw()
    root.deiconify()
    root.mainloop()


def bback2():
    windowsQ2.withdraw()
    windowsQ1.deiconify()
    windowsQ1.mainloop()


def bback3():
    windowsQ3.withdraw()
    windowsQ2.deiconify()
    windowsQ2.mainloop()


def bback4():
    windowsQ4.withdraw()
    windowsQ3.deiconify()
    windowsQ3.mainloop()


def bback5():
    windowsQ5.withdraw()
    windowsQ4.deiconify()
    windowsQ4.mainloop()


def bback6():
    windowsQ6.withdraw()
    windowsQ5.deiconify()
    windowsQ5.mainloop()


def bback7():
    windowsQ7.withdraw()
    windowsQ6.deiconify()
    windowsQ6.mainloop()


def bback8():
    windowsQ8.withdraw()
    windowsQ7.deiconify()
    windowsQ7.mainloop()


def bback9():
    windowsQ9.withdraw()
    windowsQ8.deiconify()
    windowsQ8.mainloop()


def bback10():
    windowsQ10.withdraw()
    windowsQ9.deiconify()
    windowsQ9.mainloop()


def bback11():
    windowsQ11.withdraw()
    windowsQ10.deiconify()
    windowsQ10.mainloop()


def bback12():
    windowsQ12.withdraw()
    windowsQ11.deiconify()
    windowsQ11.mainloop()


def bback13():
    windowsQ13.withdraw()
    windowsQ12.deiconify()
    windowsQ12.mainloop()


def bback14():
    windowsQ14.withdraw()
    windowsQ13.deiconify()
    windowsQ13.mainloop()


def bback15():
    windowsQ15.withdraw()
    windowsQ14.deiconify()
    windowsQ14.mainloop()


def bback16():
    windowsQ16.withdraw()
    windowsQ15.deiconify()
    windowsQ15.mainloop()


def bback17():
    windowsQ17.withdraw()
    windowsQ16.deiconify()
    windowsQ16.mainloop()


def bback18():
    windowsQ18.withdraw()
    windowsQ17.deiconify()
    windowsQ17.mainloop()


def bback19():
    windowsQ19.withdraw()
    windowsQ18.deiconify()
    windowsQ18.mainloop()


def bback20():
    windowsQ20.withdraw()
    windowsQ19.deiconify()
    windowsQ19.mainloop()




def confirmFinish():
    choice = messagebox.askyesno(title="Finishing Confirmation", message="Are you sure you want to finish this Exam?")
    if choice == True:
        finish()



def checked():

    global c
    #global w

    c=0
    d=0

    if v0.get() == 0:
        c += 1

    if v1.get() == 1:
        c += 1

    if v2.get() == 1:
        c += 1

    if v3.get() == 0:
        c += 1

    if v4.get() == 2:
        c += 1

    if v5.get() == 2:
        c += 1

    if v6.get() == 1:
        c += 1

    if v7.get() == 1:
        c += 1

    if v8.get() == 3:
        c += 1

    if v9.get() == 2:
        c += 1

    if v10.get() == 2:
        c += 1

    if v11.get() == 0:
        c += 1

    if v12.get() == 0:
        c += 1

    if v13.get() == 1:
        c += 1

    if v14.get() == 0:
        c += 1

    if v15.get() == 1:
        c += 1

    if v16.get() == 1:
        c += 1

    if v17.get() == 1:
        c += 1

    if v18.get() == 2:
        c += 1

    if v19.get() == 2:
        c += 1





def finish():
    global entry_name
    global total
    total = c
    username = entry_name.get()
    examdate = entry_date.get()
    telephoneno = entry_telephone.get()

    workbook = Workbook()
    sheet = workbook.active

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    for i in range (3,24):
        sheet.cell(row=i, column=1).border = thin_border

    for j in range (3,24):
        sheet.cell(row=j, column=2).border = thin_border

    for k in range (3,24):
        sheet.cell(row=k, column=3).border = thin_border





    sheet.column_dimensions['A'].width = 9
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 9
    sheet.column_dimensions['E'].width = 15
    sheet.column_dimensions['F'].width = 25
    sheet.column_dimensions['H'].width = 20

    sheet.cell(row=1, column=5).font = Font(size=14, bold=True)
    sheet.cell(row=2, column=1).font = Font(size=12, bold=True)
    sheet.cell(row=3, column=1).font = Font(size=12, bold=True)
    sheet.cell(row=3, column=2).font = Font(size=12, bold=True)
    sheet.cell(row=3, column=3).font = Font(size=12, bold=True)
    sheet.cell(row=1, column=3).font = Font(size=14, bold=True)


    sheet.cell(row=9, column=5).font = Font(size=14, bold=True, color="3368FF")
    sheet.cell(row=9, column=6).font = Font(size=14, bold=True, color="3368FF")
    sheet.cell(row=10, column=5).font = Font(size=14, bold=True, color="3368FF")
    sheet.cell(row=10, column=6).font = Font(size=14, bold=True, color="3368FF")
    sheet.cell(row=11, column=5).font = Font(size=14, bold=True, color="3368FF")
    sheet.cell(row=11, column=6).font = Font(size=14, bold=True, color="3368FF")
    sheet.cell(row=12, column=5).font = Font(size=14, bold=True, color="3368FF")
    sheet.cell(row=12, column=6).font = Font(size=14, bold=True, color="3368FF")
    sheet.cell(row=13, column=5).font = Font(size=14, bold=True, color="3368FF")
    sheet.cell(row=13, column=6).font = Font(size=14, bold=True, color="3368FF")

    sheet.cell(row=9, column=5).border = thin_border
    sheet.cell(row=9, column=6).border = thin_border
    sheet.cell(row=10, column=5).border = thin_border
    sheet.cell(row=10, column=6).border = thin_border
    sheet.cell(row=11, column=5).border = thin_border
    sheet.cell(row=11, column=6).border = thin_border
    sheet.cell(row=12, column=5).border = thin_border
    sheet.cell(row=12, column=6).border = thin_border
    sheet.cell(row=13, column=5).border = thin_border
    sheet.cell(row=13, column=6).border = thin_border



    sheet['C1'] = "Multiple Choice Quiz Solution"

    sheet['A3'] = "Question"
    sheet['A4'] = "1."
    sheet['A5'] = "2."
    sheet['A6'] = "3."
    sheet['A7'] = "4."
    sheet['A8'] = "5."
    sheet['A9'] = "6."
    sheet['A10'] = "7."
    sheet['A11'] = "8."
    sheet['A12'] = "9."
    sheet['A13'] = "10."
    sheet['A14'] = "11."
    sheet['A15'] = "12."
    sheet['A16'] = "13."
    sheet['A17'] = "14."
    sheet['A18'] = "15."
    sheet['A19'] = "16."
    sheet['A20'] = "17."
    sheet['A21'] = "18."
    sheet['A22'] = "19."
    sheet['A23'] = "20."

    sheet['B3'] = "Correct Answer"
    sheet['B4'] = "A"
    sheet['B5'] = "B"
    sheet['B6'] = "B"
    sheet['B7'] = "A"
    sheet['B8'] = "C"
    sheet['B9'] = "C"
    sheet['B10'] = "B"
    sheet['B11'] = "B"
    sheet['B12'] = "D"
    sheet['B13'] = "C"
    sheet['B14'] = "C"
    sheet['B15'] = "A"
    sheet['B16'] = "A"
    sheet['B17'] = "B"
    sheet['B18'] = "A"
    sheet['B19'] = "B"
    sheet['B20'] = "B"
    sheet['B21'] = "B"
    sheet['B22'] = "C"
    sheet['B23'] = "C"

    sheet['C3'] = "Result"

    if v0.get() == 0:
        sheet['C4'] = "Yes"
    else:
        sheet['C4'] = "No"

    if v1.get() == 1:
        sheet['C5'] = "Yes"
    else:
        sheet['C5'] = "No"

    if v2.get() == 1:
        sheet['C6'] = "Yes"
    else:
        sheet['C6'] = "No"

    if v3.get() == 0:
        sheet['C7'] = "Yes"
    else:
        sheet['C7'] = "No"

    if v4.get() == 2:
        sheet['C8'] = "Yes"
    else:
        sheet['C8'] = "No"

    if v5.get() == 2:
        sheet['C9'] = "Yes"
    else:
        sheet['C9'] = "No"

    if v6.get() == 1:
        sheet['C10'] = "Yes"
    else:
        sheet['C10'] = "No"

    if v7.get() == 1:
        sheet['C11'] = "Yes"
    else:
        sheet['C11'] = "No"

    if v8.get() == 3:
        sheet['C12'] = "Yes"
    else:
        sheet['C12'] = "No"

    if v9.get() == 2:
        sheet['C13'] = "Yes"
    else:
        sheet['C13'] = "No"

    if v10.get() == 2:
        sheet['C14'] = "Yes"
    else:
        sheet['C14'] = "No"

    if v11.get() == 0:
        sheet['C15'] = "Yes"
    else:
        sheet['C15'] = "No"

    if v12.get() == 0:
        sheet['C16'] = "Yes"
    else:
        sheet['C16'] = "No"

    if v13.get() == 1:
        sheet['C17'] = "Yes"
    else:
        sheet['C17'] = "No"

    if v14.get() == 0:
        sheet['C18'] = "Yes"
    else:
        sheet['C18'] = "No"

    if v15.get() == 1:
        sheet['C19'] = "Yes"
    else:
        sheet['C19'] = "No"

    if v16.get() == 1:
        sheet['C20'] = "Yes"
    else:
        sheet['C20'] = "No"

    if v17.get() == 1:
        sheet['C21'] = "Yes"
    else:
        sheet['C21'] = "No"

    if v18.get() == 2:
        sheet['C22'] = "Yes"
    else:
        sheet['C22'] = "No"

    if v19.get() == 2:
        sheet['C23'] = "Yes"
    else:
        sheet['C23'] = "No"

    global total_str
    total_str = str(total)
    examdate_str = str(examdate)
    telephoneno_str = str(telephoneno)

    sheet['C29'] = total_str

    sheet['E9'] = "Name:"
    sheet['E10'] = "Telephone:"
    sheet['E11'] = "Date:"
    sheet['E12'] = "Score:"
    sheet['E13'] = "Result:"


    sheet['F9'] = username
    sheet['F10'] = telephoneno_str
    sheet['F11'] = examdate_str
    sheet['F12'] = total_str

    if total >= 12:
        sheet['F13'] = "Passed!"
    else:
        sheet['F13'] = "Failed!"




    ###########################

    # print(entry_name.get())
    path = "exam_results"
    filename1 = os.path.join(path, username)
    workbook.save(filename=filename1 + ".xlsx")
    filename_withex = filename1 + ".xlsx"

    print_file(filename_withex)


def print_file(file_name1):
   #os.startfile(file_name1, "print")
   bnext21()


def bcheck():
    if len(entry_name.get()) == 0:
        messagebox.showwarning("Warning", "Please fill in all required fields!")
    elif len(entry_telephone.get()) == 0:
        messagebox.showwarning("Warning", "Please fill in all required fields!")
    elif len(entry_date.get()) == 0:
        messagebox.showwarning("Warning", "Please fill in all required fields!")
    else:
        bnext()


root = tkinter.Tk()

v0 = IntVar()
v0.set(-1)
v1 = IntVar()
v1.set(-1)
v2 = IntVar()
v2.set(-1)
v3 = IntVar()
v3.set(-1)
v4 = IntVar()
v4.set(-1)
v5 = IntVar()
v5.set(-1)
v6 = IntVar()
v6.set(-1)
v7 = IntVar()
v7.set(-1)
v8 = IntVar()
v8.set(-1)
v9 = IntVar()
v9.set(-1)
v10 = IntVar()
v10.set(-1)
v11 = IntVar()
v11.set(-1)
v12 = IntVar()
v12.set(-1)
v13 = IntVar()
v13.set(-1)
v14 = IntVar()
v14.set(-1)
v15 = IntVar()
v15.set(-1)
v16 = IntVar()
v16.set(-1)
v17 = IntVar()
v17.set(-1)
v18 = IntVar()
v18.set(-1)
v19 = IntVar()
v19.set(-1)



root.title("Multiple Choice Quiz")
root.geometry("1920x1080")
root.resizable(0, 0)



load = image1.open("logo.png")
load.thumbnail((150,150))
render = ImageTk.PhotoImage(load)

# labels can be text or images
img = Label(root, image=render)
img.image = render
img.place(x=865, y=30)



label_title = Label(root, text="Multiple Choice Quiz", width=55, font=("bold", 24))
label_title.place(x=420, y=230)

label_name = Label(root, text="Full Name:", font=("bold", 20))
label_name.place(x=530, y=380)

entry_name = Entry(root, font=("bold", 20), width=40)
entry_name.place(x=670, y=380)


label_telephone = Label(root, text="Telephone:", font=("bold", 20))
label_telephone.place(x=530, y=480)

entry_telephone = Entry(root, font=("bold", 20), width=40)
entry_telephone.place(x=670, y=480)


label_date = Label(root, text="Date:", font=("bold", 20))
label_date.place(x=530, y=580)

entry_date = Entry(root, font=("bold", 20), width=40)
entry_date.place(x=670, y=580)


btn1 = Button(root,text = "Start",font = ('arial',18,'bold'),fg = 'blue',command = bcheck, height = 1, width = 5).pack(side=BOTTOM, pady=280)

root.mainloop()